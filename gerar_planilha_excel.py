#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Checklist Desenvolvimento"

# ============ CONFIGURAÇÕES GERAIS ============
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Azul
header_font = Font(bold=True, color="FFFFFF", size=12)

subheader_fill = PatternFill(start_color="6633CC", end_color="6633CC", fill_type="solid")  # Roxo
subheader_font = Font(bold=True, color="FFFFFF", size=11)

completed_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde
pending_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Amarelo
error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Vermelho

# ============ CABEÇALHO ============
ws['A1'] = "🚀 APOGEU - CHECKLIST DE DESENVOLVIMENTO"
ws['A1'].font = Font(bold=True, size=16, color="0066CC")
ws.merge_cells('A1:F1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws['A2'] = "Seu Pico de Sucesso em Marketing Digital"
ws['A2'].font = Font(italic=True, size=11, color="6633CC")
ws.merge_cells('A2:F2')
ws['A2'].alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 20

# ============ SEÇÃO 1: CHECKLIST DE INSTALAÇÃO ============
ws['A4'] = "FASE 1: INSTALAÇÃO E CONFIGURAÇÃO"
ws['A4'].font = subheader_font
ws['A4'].fill = subheader_fill
ws.merge_cells('A4:F4')
ws['A4'].alignment = Alignment(horizontal='center')
ws.row_dimensions[4].height = 25

# Headers da tabela
headers = ['#', 'Tarefa', 'Status', 'Data Conclusão', 'Responsável', 'Observações']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

ws.row_dimensions[5].height = 20

# Dados da instalação
tasks = [
    ('1', 'Verificar Windows 10 Pro', 'Pendente', '', '', ''),
    ('2', 'Desativar Antivírus', 'Pendente', '', '', ''),
    ('3', 'Instalar Node.js 18+', 'Pendente', '', '', ''),
    ('4', 'Instalar Git', 'Pendente', '', '', ''),
    ('5', 'Instalar MySQL 8.0', 'Pendente', '', '', ''),
    ('6', 'Instalar pnpm', 'Pendente', '', '', ''),
    ('7', 'Criar pasta C:\\Projetos', 'Pendente', '', '', ''),
    ('8', 'Clonar repositório GitHub', 'Pendente', '', '', ''),
    ('9', 'Instalar dependências (pnpm install)', 'Pendente', '', '', ''),
    ('10', 'Criar banco de dados MySQL', 'Pendente', '', '', ''),
    ('11', 'Configurar arquivo .env', 'Pendente', '', '', ''),
    ('12', 'Executar migrações (pnpm db:push)', 'Pendente', '', '', ''),
    ('13', 'Iniciar Backend (pnpm dev)', 'Pendente', '', '', ''),
    ('14', 'Iniciar Frontend (pnpm dev)', 'Pendente', '', '', ''),
    ('15', 'Acessar http://localhost:5173', 'Pendente', '', '', ''),
    ('16', 'Fazer login na aplicação', 'Pendente', '', '', ''),
    ('17', 'Executar testes (pnpm test)', 'Pendente', '', '', ''),
    ('18', 'Verificar TypeScript (pnpm check)', 'Pendente', '', '', ''),
    ('19', 'Fazer Build (pnpm build)', 'Pendente', '', '', ''),
    ('20', 'Criar App Desktop', 'Pendente', '', '', ''),
]

row = 6
for task in tasks:
    for col, value in enumerate(task, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Colorir coluna de status
        if col == 3:  # Coluna Status
            if value == 'Concluído':
                cell.fill = completed_fill
            elif value == 'Em Progresso':
                cell.fill = pending_fill
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    row += 1

# ============ SEÇÃO 2: PROGRESSO GERAL ============
ws['A28'] = "FASE 2: PROGRESSO GERAL"
ws['A28'].font = subheader_font
ws['A28'].fill = subheader_fill
ws.merge_cells('A28:F28')
ws['A28'].alignment = Alignment(horizontal='center')
ws.row_dimensions[28].height = 25

# Tabela de progresso
progress_headers = ['Métrica', 'Valor', 'Percentual', 'Status']
for col, header in enumerate(progress_headers, 1):
    cell = ws.cell(row=29, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

ws.row_dimensions[29].height = 20

progress_data = [
    ('Tarefas Concluídas', 0, '0%', 'Iniciando'),
    ('Tarefas em Progresso', 0, '0%', 'Aguardando'),
    ('Tarefas Pendentes', 20, '100%', 'Não Iniciado'),
    ('Taxa de Conclusão', 0, '0%', 'Acompanhamento'),
]

for idx, (metric, value, percent, status) in enumerate(progress_data, 30):
    ws.cell(row=idx, column=1).value = metric
    ws.cell(row=idx, column=2).value = value
    ws.cell(row=idx, column=3).value = percent
    ws.cell(row=idx, column=4).value = status
    
    for col in range(1, 5):
        cell = ws.cell(row=idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if col == 4:
            if status == 'Concluído':
                cell.fill = completed_fill
            elif status == 'Aguardando':
                cell.fill = pending_fill
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# ============ SEÇÃO 3: RASTREAMENTO DE TEMPO ============
ws['A36'] = "FASE 3: RASTREAMENTO DE TEMPO"
ws['A36'].font = subheader_font
ws['A36'].fill = subheader_fill
ws.merge_cells('A36:F36')
ws['A36'].alignment = Alignment(horizontal='center')
ws.row_dimensions[36].height = 25

time_headers = ['Etapa', 'Tempo Estimado', 'Tempo Real', 'Diferença', 'Status']
for col, header in enumerate(time_headers, 1):
    cell = ws.cell(row=37, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

ws.row_dimensions[37].height = 20

time_data = [
    ('Instalação Node.js', '10 min', '', '', 'Pendente'),
    ('Instalação Git', '5 min', '', '', 'Pendente'),
    ('Instalação MySQL', '15 min', '', '', 'Pendente'),
    ('Instalação pnpm', '3 min', '', '', 'Pendente'),
    ('Download e Setup', '20 min', '', '', 'Pendente'),
    ('Configuração Banco', '10 min', '', '', 'Pendente'),
    ('Instalação Dependências', '15 min', '', '', 'Pendente'),
    ('Execução Migrações', '5 min', '', '', 'Pendente'),
    ('Testes e Verificação', '10 min', '', '', 'Pendente'),
]

for idx, (stage, estimated, real, diff, status) in enumerate(time_data, 38):
    ws.cell(row=idx, column=1).value = stage
    ws.cell(row=idx, column=2).value = estimated
    ws.cell(row=idx, column=3).value = real
    ws.cell(row=idx, column=4).value = diff
    ws.cell(row=idx, column=5).value = status
    
    for col in range(1, 6):
        cell = ws.cell(row=idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

# ============ SEÇÃO 4: ERROS E SOLUÇÕES ============
ws['A49'] = "FASE 4: ERROS E SOLUÇÕES"
ws['A49'].font = subheader_font
ws['A49'].fill = subheader_fill
ws.merge_cells('A49:F49')
ws['A49'].alignment = Alignment(horizontal='center')
ws.row_dimensions[49].height = 25

error_headers = ['Erro', 'Descrição', 'Solução', 'Resolvido?', 'Data', 'Notas']
for col, header in enumerate(error_headers, 1):
    cell = ws.cell(row=50, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

ws.row_dimensions[50].height = 20

error_data = [
    ('', '', '', 'Não', '', 'Registre aqui qualquer erro encontrado'),
    ('', '', '', 'Não', '', ''),
    ('', '', '', 'Não', '', ''),
    ('', '', '', 'Não', '', ''),
    ('', '', '', 'Não', '', ''),
]

for idx, (error, desc, solution, resolved, date, notes) in enumerate(error_data, 51):
    ws.cell(row=idx, column=1).value = error
    ws.cell(row=idx, column=2).value = desc
    ws.cell(row=idx, column=3).value = solution
    ws.cell(row=idx, column=4).value = resolved
    ws.cell(row=idx, column=5).value = date
    ws.cell(row=idx, column=6).value = notes
    
    for col in range(1, 7):
        cell = ws.cell(row=idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

# ============ SEÇÃO 5: CONFIGURAÇÕES DE API ============
ws['A59'] = "FASE 5: CONFIGURAÇÕES DE API (OPCIONAL)"
ws['A59'].font = subheader_font
ws['A59'].fill = subheader_fill
ws.merge_cells('A59:F59')
ws['A59'].alignment = Alignment(horizontal='center')
ws.row_dimensions[59].height = 25

api_headers = ['API', 'Configurada?', 'Chave', 'Data Config.', 'Testada?', 'Observações']
for col, header in enumerate(api_headers, 1):
    cell = ws.cell(row=60, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

ws.row_dimensions[60].height = 20

api_data = [
    ('Google Ads', 'Não', '', '', 'Não', ''),
    ('Meta Ads', 'Não', '', '', 'Não', ''),
    ('TikTok Ads', 'Não', '', '', 'Não', ''),
    ('OpenAI', 'Não', '', '', 'Não', ''),
    ('ElevenLabs', 'Não', '', '', 'Não', ''),
    ('Twilio (WhatsApp)', 'Não', '', '', 'Não', ''),
    ('PagBrasil', 'Não', '', '', 'Não', ''),
]

for idx, (api, configured, key, date, tested, notes) in enumerate(api_data, 61):
    ws.cell(row=idx, column=1).value = api
    ws.cell(row=idx, column=2).value = configured
    ws.cell(row=idx, column=3).value = key
    ws.cell(row=idx, column=4).value = date
    ws.cell(row=idx, column=5).value = tested
    ws.cell(row=idx, column=6).value = notes
    
    for col in range(1, 7):
        cell = ws.cell(row=idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left')

# ============ AJUSTAR LARGURA DAS COLUNAS ============
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 30

# ============ CRIAR SEGUNDA ABA - GRÁFICOS ============
ws_charts = wb.create_sheet("Gráficos e Análises")

# Título
ws_charts['A1'] = "📊 GRÁFICOS E ANÁLISES DE PROGRESSO"
ws_charts['A1'].font = Font(bold=True, size=16, color="0066CC")
ws_charts.merge_cells('A1:D1')
ws_charts['A1'].alignment = Alignment(horizontal='center')
ws_charts.row_dimensions[1].height = 30

# Dados para gráficos
ws_charts['A3'] = "Status do Projeto"
ws_charts['A3'].font = Font(bold=True, size=12, color="0066CC")

status_data = [
    ['Status', 'Quantidade'],
    ['Concluído', 0],
    ['Em Progresso', 0],
    ['Pendente', 20],
]

for row_idx, row_data in enumerate(status_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        ws_charts.cell(row=row_idx, column=col_idx).value = value
        if row_idx == 3:
            ws_charts.cell(row=row_idx, column=col_idx).font = Font(bold=True)

# Gráfico de Pizza
pie = PieChart()
pie.title = "Distribuição de Tarefas"
pie.style = 10
labels = Reference(ws_charts, min_col=1, min_row=4, max_row=6)
data = Reference(ws_charts, min_col=2, min_row=3, max_row=6)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
ws_charts.add_chart(pie, "A8")

# Dados para gráfico de barras
ws_charts['E3'] = "Progresso por Fase"
ws_charts['E3'].font = Font(bold=True, size=12, color="0066CC")

phase_data = [
    ['Fase', 'Progresso %'],
    ['Instalação', 0],
    ['Configuração', 0],
    ['Desenvolvimento', 0],
    ['Testes', 0],
    ['Deploy', 0],
]

for row_idx, row_data in enumerate(phase_data, 3):
    for col_idx, value in enumerate(row_data, 5):
        ws_charts.cell(row=row_idx, column=col_idx).value = value
        if row_idx == 3:
            ws_charts.cell(row=row_idx, column=col_idx).font = Font(bold=True)

# Gráfico de Barras
bar = BarChart()
bar.type = "col"
bar.title = "Progresso por Fase (%)"
bar.style = 10
bar.y_axis.title = 'Progresso (%)'
bar.x_axis.title = 'Fases'
data_bar = Reference(ws_charts, min_col=6, min_row=3, max_row=8)
cats = Reference(ws_charts, min_col=5, min_row=4, max_row=8)
bar.add_data(data_bar, titles_from_data=True)
bar.set_categories(cats)
ws_charts.add_chart(bar, "E8")

# Dados para gráfico de linha
ws_charts['A25'] = "Cronograma de Execução"
ws_charts['A25'].font = Font(bold=True, size=12, color="0066CC")

timeline_data = [
    ['Semana', 'Tarefas Concluídas'],
    ['Semana 1', 0],
    ['Semana 2', 0],
    ['Semana 3', 0],
    ['Semana 4', 0],
]

for row_idx, row_data in enumerate(timeline_data, 25):
    for col_idx, value in enumerate(row_data, 1):
        ws_charts.cell(row=row_idx, column=col_idx).value = value
        if row_idx == 25:
            ws_charts.cell(row=row_idx, column=col_idx).font = Font(bold=True)

# Gráfico de Linha
line = LineChart()
line.title = "Progresso Semanal"
line.style = 10
line.y_axis.title = 'Tarefas Concluídas'
line.x_axis.title = 'Semanas'
data_line = Reference(ws_charts, min_col=2, min_row=25, max_row=29)
cats_line = Reference(ws_charts, min_col=1, min_row=26, max_row=29)
line.add_data(data_line, titles_from_data=True)
line.set_categories(cats_line)
ws_charts.add_chart(line, "A30")

# Ajustar largura
ws_charts.column_dimensions['A'].width = 25
ws_charts.column_dimensions['B'].width = 25
ws_charts.column_dimensions['E'].width = 25
ws_charts.column_dimensions['F'].width = 25

# ============ CRIAR TERCEIRA ABA - NOTAS ============
ws_notes = wb.create_sheet("Notas e Observações")

ws_notes['A1'] = "📝 NOTAS E OBSERVAÇÕES"
ws_notes['A1'].font = Font(bold=True, size=16, color="0066CC")
ws_notes.merge_cells('A1:D1')
ws_notes['A1'].alignment = Alignment(horizontal='center')
ws_notes.row_dimensions[1].height = 30

ws_notes['A3'] = "Use esta aba para registrar observações, dúvidas e aprendizados durante a instalação."
ws_notes['A3'].font = Font(italic=True, size=11)
ws_notes.merge_cells('A3:D3')

ws_notes['A5'] = "Data"
ws_notes['B5'] = "Observação"
for col in ['A', 'B']:
    ws_notes[f'{col}5'].font = header_font
    ws_notes[f'{col}5'].fill = header_fill
    ws_notes[f'{col}5'].border = thin_border
    ws_notes[f'{col}5'].alignment = Alignment(horizontal='center')

ws_notes.column_dimensions['A'].width = 20
ws_notes.column_dimensions['B'].width = 80

for row in range(6, 30):
    for col in ['A', 'B']:
        cell = ws_notes[f'{col}{row}']
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
    ws_notes.row_dimensions[row].height = 30

# ============ SALVAR WORKBOOK ============
wb.save('/home/ubuntu/apogeu/CONTROLE_DESENVOLVIMENTO_APOGEU.xlsx')
print('✅ Planilha Excel criada com sucesso!')
print('📊 Arquivo: CONTROLE_DESENVOLVIMENTO_APOGEU.xlsx')
print('   - Aba 1: Checklist Desenvolvimento')
print('   - Aba 2: Gráficos e Análises')
print('   - Aba 3: Notas e Observações')

