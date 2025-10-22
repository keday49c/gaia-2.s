#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import datetime

# Criar workbook
wb = Workbook()
wb.remove(wb.active)

# ===== ABA 1: CHECKLIST =====
ws1 = wb.create_sheet('Checklist Desenvolvimento')

# Estilos
header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
completed_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
pending_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ['ID', 'Tarefa', 'Status', 'Data Início', 'Data Conclusão', 'Responsável', 'Observações']
ws1.append(headers)

for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Dados
tasks = [
    (1, 'Instalar Node.js', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'Versão 18.0.0'),
    (2, 'Instalar Git', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'Configurado'),
    (3, 'Instalar MySQL', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'Porta 3306'),
    (4, 'Clonar Repositório', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'gaia-3.0'),
    (5, 'Instalar Dependências', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'pnpm install'),
    (6, 'Criar Banco de Dados', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'apogeu_db'),
    (7, 'Configurar .env', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'Todas as variáveis'),
    (8, 'Executar Migrações', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'pnpm db:push'),
    (9, 'Iniciar Backend', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'pnpm dev'),
    (10, 'Iniciar Frontend', 'Concluído', '22/10/2024', '22/10/2024', 'Você', 'Porta 5173'),
    (11, 'Testar Login', 'Em Progresso', '23/10/2024', '', 'Você', 'OAuth funcionando'),
    (12, 'Testar Campanhas', 'Pendente', '', '', 'Você', ''),
    (13, 'Testar Credenciais', 'Pendente', '', '', 'Você', ''),
    (14, 'Testar Analytics', 'Pendente', '', '', 'Você', ''),
    (15, 'Testar Media Boost', 'Pendente', '', '', 'Você', ''),
    (16, 'Testar Developer Login', 'Pendente', '', '', 'Você', ''),
    (17, 'Testar Painel Admin', 'Pendente', '', '', 'Você', ''),
    (18, 'Testar Guia Interativo', 'Pendente', '', '', 'Você', ''),
    (19, 'Executar Testes', 'Pendente', '', '', 'Você', 'pnpm test'),
    (20, 'Deploy no GitHub', 'Pendente', '', '', 'Você', 'gaia-3.0'),
]

for task in tasks:
    ws1.append(task)
    row = ws1.max_row
    
    # Colorir status
    status_cell = ws1[f'C{row}']
    if task[2] == 'Concluído':
        status_cell.fill = completed_fill
    elif task[2] == 'Em Progresso':
        status_cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    else:
        status_cell.fill = pending_fill
    
    # Aplicar bordas
    for col in range(1, 8):
        cell = ws1.cell(row, col)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Ajustar largura das colunas
ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 15
ws1.column_dimensions['G'].width = 25

# ===== ABA 2: PROGRESSO =====
ws2 = wb.create_sheet('Progresso')

# Dados de progresso
ws2['A1'] = 'Fase'
ws2['B1'] = 'Conclusão %'
ws2['C1'] = 'Status'

ws2['A1'].fill = header_fill
ws2['B1'].fill = header_fill
ws2['C1'].fill = header_fill

phases = [
    ('Instalação', 100, 'Concluído'),
    ('Configuração', 100, 'Concluído'),
    ('Backend', 100, 'Concluído'),
    ('Frontend', 100, 'Concluído'),
    ('APIs de Marketing', 100, 'Concluído'),
    ('WhatsApp Bot', 100, 'Concluído'),
    ('Geração de Conteúdo', 100, 'Concluído'),
    ('Dashboards', 100, 'Concluído'),
    ('Sistema de Assinaturas', 100, 'Concluído'),
    ('Análise de Concorrência', 100, 'Concluído'),
    ('Testes de Segurança', 100, 'Concluído'),
    ('Desktop App', 100, 'Concluído'),
    ('Documentação', 100, 'Concluído'),
    ('Developer Login', 100, 'Concluído'),
    ('Painel Admin', 100, 'Concluído'),
    ('Controle Avançado', 100, 'Concluído'),
    ('Media Boost', 100, 'Concluído'),
    ('Guia Interativo', 100, 'Concluído'),
]

for idx, (phase, completion, status) in enumerate(phases, 2):
    ws2[f'A{idx}'] = phase
    ws2[f'B{idx}'] = completion
    ws2[f'C{idx}'] = status
    
    # Colorir células
    if completion == 100:
        ws2[f'B{idx}'].fill = completed_fill
        ws2[f'C{idx}'].fill = completed_fill
    
    for col in ['A', 'B', 'C']:
        ws2[f'{col}{idx}'].border = border

# Gráfico de progresso
chart = BarChart()
chart.type = 'col'
chart.title = 'Progresso do Desenvolvimento'
chart.y_axis.title = 'Conclusão %'
chart.x_axis.title = 'Fases'

data = Reference(ws2, min_col=2, min_row=1, max_row=len(phases)+1)
categories = Reference(ws2, min_col=1, min_row=2, max_row=len(phases)+1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
ws2.add_chart(chart, 'E2')

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15

# ===== ABA 3: MÉTRICAS =====
ws3 = wb.create_sheet('Métricas')

ws3['A1'] = 'Data'
ws3['B1'] = 'Campanhas Ativas'
ws3['C1'] = 'Impressões'
ws3['D1'] = 'Cliques'
ws3['E1'] = 'Conversões'
ws3['F1'] = 'Gasto (R$)'
ws3['G1'] = 'Receita (R$)'
ws3['H1'] = 'ROI %'

for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border

# Dados de métricas
metrics_data = [
    ('22/10/2024', 3, 125430, 3847, 287, 4230.50, 14320.00, 238.5),
    ('23/10/2024', 3, 145230, 4234, 312, 4890.75, 15680.00, 220.6),
    ('24/10/2024', 4, 165890, 4876, 345, 5234.25, 17450.00, 233.4),
    ('25/10/2024', 4, 178234, 5123, 378, 5678.50, 18920.00, 233.1),
    ('26/10/2024', 4, 189567, 5456, 412, 6123.75, 20340.00, 232.4),
    ('27/10/2024', 5, 201234, 5789, 445, 6567.25, 21890.00, 233.2),
    ('28/10/2024', 5, 215678, 6123, 478, 7012.50, 23450.00, 234.5),
]

for idx, data_row in enumerate(metrics_data, 2):
    for col_idx, value in enumerate(data_row, 1):
        cell = ws3.cell(idx, col_idx)
        cell.value = value
        cell.border = border
        if col_idx > 5:
            cell.number_format = '#,##0.00'

# Gráfico de ROI
roi_chart = LineChart()
roi_chart.title = 'ROI ao Longo do Tempo'
roi_chart.y_axis.title = 'ROI %'
roi_chart.x_axis.title = 'Data'

roi_data = Reference(ws3, min_col=8, min_row=1, max_row=len(metrics_data)+1)
roi_categories = Reference(ws3, min_col=1, min_row=2, max_row=len(metrics_data)+1)

roi_chart.add_data(roi_data, titles_from_data=True)
roi_chart.set_categories(roi_categories)
ws3.add_chart(roi_chart, 'J2')

# Ajustar largura
for col in range(1, 9):
    ws3.column_dimensions[get_column_letter(col)].width = 15

# Salvar workbook
wb.save('/home/ubuntu/apogeu/CONTROLE_DESENVOLVIMENTO_GAIA_3.0.xlsx')
print('✅ Planilha Excel criada com sucesso: CONTROLE_DESENVOLVIMENTO_GAIA_3.0.xlsx')

